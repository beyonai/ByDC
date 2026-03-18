"""生成《知识构建模板.xlsx》导入模板文件。

Sheet 顺序：领域 → 知识库 → 术语类型 → 术语 → 术语关系 → 📁说明

运行：
    python packages/datacloud-knowledge/docs/模块设计/_gen_import_template.py
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── 颜色常量 ──────────────────────────────────────────────────────────────────
CLR_HDR_REQ  = "1F497D"   # 深蓝：必填列标题
CLR_HDR_OPT  = "4472C4"   # 中蓝：可选列标题
CLR_HDR_FONT = "FFFFFF"   # 白色字体
CLR_EX_ODD   = "EBF3FB"   # 淡蓝：奇数示例行
CLR_EX_EVEN  = "DEEAF1"   # 稍深蓝：偶数示例行
CLR_NOTICE   = "FFF2CC"   # 淡黄：说明行底色
CLR_NTC_FONT = "7F6000"   # 棕黄：说明行字体

_thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

# 大分类中文值（用于下拉）
CATEGORY_OPTS = "列表术语,字典术语,本体术语,文档名称术语"


# ── 通用样式工具 ──────────────────────────────────────────────────────────────
def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _style_hdr(cell, required: bool = True) -> None:
    cell.fill      = _fill(CLR_HDR_REQ if required else CLR_HDR_OPT)
    cell.font      = Font(bold=True, color=CLR_HDR_FONT, size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = BORDER


def _style_ex(cell, row_offset: int = 0) -> None:
    cell.fill      = _fill(CLR_EX_ODD if row_offset % 2 == 0 else CLR_EX_EVEN)
    cell.font      = Font(size=10)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border    = BORDER


def _style_notice(cell) -> None:
    cell.fill      = _fill(CLR_NOTICE)
    cell.font      = Font(color=CLR_NTC_FONT, italic=True, size=9)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border    = BORDER


def _col_w(ws, col: int, w: float) -> None:
    ws.column_dimensions[get_column_letter(col)].width = w


def _notice_row(ws, row: int, text: str, n_cols: int, height: int = 30) -> None:
    """合并整行写说明文字。"""
    ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
    cell = ws[f"A{row}"]
    cell.value = text
    _style_notice(cell)
    ws.row_dimensions[row].height = height


def _header_row(ws, row: int, cols: list) -> None:
    """写表头行（每列 (display, required, width)）。"""
    for i, (display, required, width) in enumerate(cols, 1):
        cell = ws.cell(row=row, column=i, value=display)
        _style_hdr(cell, required)
        _col_w(ws, i, width)
    ws.row_dimensions[row].height = 36


def _desc_row(ws, row: int, descs: list[str], height: int = 40) -> None:
    """写字段说明行。"""
    for i, desc in enumerate(descs, 1):
        cell = ws.cell(row=row, column=i, value=desc)
        _style_notice(cell)
    ws.row_dimensions[row].height = height


def _example_rows(ws, start_row: int, examples: list[tuple]) -> None:
    for offset, row_data in enumerate(examples):
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=start_row + offset, column=col, value=val)
            _style_ex(cell, offset)


def _dv_list(ws, formula: str, cells: str, allow_blank: bool = True,
             err_msg: str = "", err_title: str = "格式错误") -> DataValidation:
    dv = DataValidation(type="list", formula1=formula, allow_blank=allow_blank)
    if err_msg:
        dv.error      = err_msg
        dv.errorTitle = err_title
    ws.add_data_validation(dv)
    dv.add(cells)
    return dv


# ─────────────────────────────────────────────────────────────────────────────
# Sheet1：领域
# ─────────────────────────────────────────────────────────────────────────────
_DOMAIN_COLS = [
    # (display, required, width, desc)
    ("领域编码 *",  True,  20, "唯一标识，全英文+下划线，如 sales / hr / finance"),
    ("领域名称 *",  True,  22, "中文名称，如 销售领域 / 人力资源"),
    ("父领域编码",  False, 20, "填父级的【领域编码】；根节点留空"),
    ("领域描述",    False, 35, "可选，对该领域的业务说明"),
]

_DOMAIN_EXAMPLES = [
    ("sales",    "销售领域",   "",       "覆盖CRM、商机、客户等销售业务"),
    ("hr",       "人力资源",   "",       "员工、组织、KPI 等人事相关"),
    ("sales_crm","CRM子领域",  "sales",  ""),
]


def _build_domain(wb: Workbook) -> None:
    ws = wb.create_sheet("领域")
    n = len(_DOMAIN_COLS)
    _notice_row(ws, 1,
        "【领域】填写业务领域分类。领域编码将在《术语》Sheet 中用于下拉选择。"
        "支持树形层级（通过父领域编码关联），根节点父领域编码留空。",
        n, height=28)
    _header_row(ws, 2, [(d, r, w) for d, r, w, _ in _DOMAIN_COLS])
    _desc_row(ws, 3, [desc for _, _, _, desc in _DOMAIN_COLS])
    _example_rows(ws, 4, _DOMAIN_EXAMPLES)
    ws.freeze_panes = "A4"


# ─────────────────────────────────────────────────────────────────────────────
# Sheet2：知识库
# ─────────────────────────────────────────────────────────────────────────────
_LIBRARY_COLS = [
    ("知识库编码 *", True,  22, "唯一标识，全英文+下划线，如 crm_kb / hr_kb"),
    ("知识库名称 *", True,  28, "中文名称，如 CRM知识库 / 销售培训知识库"),
]

_LIBRARY_EXAMPLES = [
    ("crm_kb",    "CRM知识库"),
    ("hr_kb",     "人力资源知识库"),
    ("product_kb","产品知识库"),
]


def _build_library(wb: Workbook) -> None:
    ws = wb.create_sheet("知识库")
    n = len(_LIBRARY_COLS)
    _notice_row(ws, 1,
        "【知识库】定义术语的来源归属。知识库编码将在《术语》Sheet 中用于下拉选择。",
        n, height=24)
    _header_row(ws, 2, [(d, r, w) for d, r, w, _ in _LIBRARY_COLS])
    _desc_row(ws, 3, [desc for _, _, _, desc in _LIBRARY_COLS])
    _example_rows(ws, 4, _LIBRARY_EXAMPLES)
    ws.freeze_panes = "A4"


# ─────────────────────────────────────────────────────────────────────────────
# Sheet3：术语类型
# ─────────────────────────────────────────────────────────────────────────────
_TYPE_COLS = [
    ("类型编码 *",  True,  18, "唯一标识，大写英文，如 OBJ / VIEW / ACTION"),
    ("类型名称 *",  True,  18, "中文显示名，如 对象 / 视图 / 动作"),
    ("类型描述",    False, 32, "对该类型的业务说明，可留空"),
    ("大分类 *",    True,  18, "从下拉选择：列表术语 / 字典术语 / 本体术语 / 文档名称术语"),
    ("是否内置",    False, 14, "true=系统预置不可删  false=用户自定义（默认 false）"),
]

_TYPE_EXAMPLES = [
    ("OBJ",           "对象",     "本体-对象类型，如客户、合同",      "本体术语",   "false"),
    ("VIEW",          "视图",     "本体-视图类型，如销售分析视图",     "本体术语",   "false"),
    ("ACTION",        "动作",     "本体-动作类型，如查询、提交",       "本体术语",   "false"),
    ("FUNC",          "函数",     "本体-函数类型，如聚合函数",         "本体术语",   "false"),
    ("EMPLOYEE",      "员工",     "员工列表术语",                       "列表术语",   "false"),
    ("CUSTOMER_TYPE", "客户类型", "客户类别字典",                       "字典术语",   "false"),
]


def _build_term_type(wb: Workbook) -> None:
    ws = wb.create_sheet("术语类型")
    n = len(_TYPE_COLS)
    _notice_row(ws, 1,
        "【术语类型】定义术语的分类编码。大分类从下拉选择中文选项，系统导入时自动转换为数字。"
        "内置类型（is_builtin=true）由系统预置，此处仅需填写用户自定义类型。",
        n, height=32)
    _header_row(ws, 2, [(d, r, w) for d, r, w, _ in _TYPE_COLS])
    _desc_row(ws, 3, [desc for _, _, _, desc in _TYPE_COLS])
    _example_rows(ws, 4, _TYPE_EXAMPLES)

    # 大分类下拉（中文）
    _dv_list(ws, f'"{CATEGORY_OPTS}"', "D4:D1000", allow_blank=False,
             err_msg="请从下拉列表中选择大分类", err_title="格式错误")
    # 是否内置下拉
    _dv_list(ws, '"true,false"', "E4:E1000")

    ws.freeze_panes = "A4"


# ─────────────────────────────────────────────────────────────────────────────
# Sheet4：术语
# ─────────────────────────────────────────────────────────────────────────────
_TERM_COLS = [
    # (display, required, width, desc)
    ("术语编码",        False, 22,
     "留空=新增（系统生成）；填写已有编码=修改该术语；导入后作为唯一标识不可更改"),
    ("术语名称 *",      True,  25,
     "标准规范名称，全局唯一，在《术语关系》中用 source/target 引用"),
    ("类型编码 *",      True,  18,
     "必须与《术语类型》Sheet 中 type_code 一致（含系统内置类型）"),
    ("领域编码",        False, 20,
     "从下拉选择《领域》Sheet 中的领域编码；可留空"),
    ("知识库编码",      False, 20,
     "从下拉选择《知识库》Sheet 中的知识库编码；可留空"),
    ("父术语编码",      False, 22,
     "仅实例术语填写：填父概念术语的【术语编码】；概念术语留空"),
    ("本体定义文件路径", False, 32,
     "相对于【ontology/】目录的子路径，如 objects/sales_customer.json\n"
     "仅本体术语（大分类=本体术语）填写；其他类型留空"),
    ("描述摘要",        False, 35,
     "100 字以内简短描述，用于搜索展示"),
    ("别名（逗号分隔）", False, 28,
     "多个别名用英文逗号分隔，如：客户,CRM客户"),
    ("标签属性（JSON）", False, 35,
     '可留空；格式：{"tag维度term_id":{"type":"list","value":"具体值term_id"}}'),
]

_TERM_EXAMPLES = [
    ("",  "销售客户对象",    "OBJ",    "sales", "crm_kb", "",       "objects/sales_customer.json",
     "CRM客户表，记录客户基本信息",  "客户,CRM客户",  ""),
    ("",  "在线查数分析场景", "VIEW",   "sales", "",       "",       "views/scene_01_data_analysis.json",
     "销售数据分析场景",              "查数场景",      ""),
    ("",  "查询待办",        "ACTION", "sales", "",       "",       "",
     "查询当前用户的待办列表",        "查待办",        ""),
    ("",  "广州分公司",      "EMPLOYEE","hr",   "hr_kb",  "组织",   "",
     "广州分公司全体员工",            "",              ""),
    ("",  "客户类型",        "CUSTOMER_TYPE","sales","",  "",       "",
     "客户所属行业类型枚举",          "",              ""),
]


def _build_term(wb: Workbook) -> None:
    ws = wb.create_sheet("术语")
    n = len(_TERM_COLS)
    _notice_row(ws, 1,
        "【术语】* 为必填项。\n"
        "① 术语编码：留空=新增，填已有编码=修改；导入后作为唯一标识不可更改。\n"
        "② 领域编码 / 知识库编码：从下拉选择，与《领域》《知识库》Sheet 联动。\n"
        "③ 本体定义文件路径：相对于【ontology/】目录的子路径，如 objects/sales_customer.json。\n"
        "④ 父术语编码：填父概念术语的【术语编码】，实例术语填写，概念术语留空。",
        n, height=65)
    _header_row(ws, 2, [(d, r, w) for d, r, w, _ in _TERM_COLS])
    _desc_row(ws, 3, [desc for _, _, _, desc in _TERM_COLS], height=50)
    _example_rows(ws, 4, _TERM_EXAMPLES)

    # 领域编码下拉（联动《领域》Sheet A列）
    _dv_list(ws, "领域!$A$4:$A$1000", "D4:D1000")
    # 知识库编码下拉（联动《知识库》Sheet A列）
    _dv_list(ws, "知识库!$A$4:$A$1000", "E4:E1000")

    ws.freeze_panes = "A4"


# ─────────────────────────────────────────────────────────────────────────────
# Sheet5：术语关系
# ─────────────────────────────────────────────────────────────────────────────
_REL_COLS = [
    ("源术语编码 *",      True,  25,
     "填《术语》Sheet 中的【术语编码】（留空时填【术语名称】）"),
    ("目标术语编码 *",    True,  25,
     "填《术语》Sheet 中的【术语编码】（留空时填【术语名称】）"),
    ("关系名称 *",        True,  30,
     "建议格式：源术语_动词_目标术语，如 销售人员_负责_客户"),
    ("关系类别 *",        True,  18,
     "ONTOLOGY=本体结构关系  BUSINESS=业务自定义关系"),
    ("数量约束",          False, 12,
     "1:1 | 1:N | N:1 | N:N；留空默认 N:N"),
    ("动作术语编码",      False, 25,
     "绑定的 ACTION 类型术语编码（或名称）；BUSINESS 关系推荐填写，ONTOLOGY 通常留空"),
]

_REL_EXAMPLES = [
    ("sales_customer_obj", "scene_01_data_analysis", "客户对象_属于_查数场景", "ONTOLOGY", "N:N", ""),
    ("sales_customer_obj", "hr_org",                 "客户_归属_组织",          "BUSINESS", "N:1", "query_todo"),
    ("scene_01",           "sales_customer_obj",     "查数场景_包含_客户对象",   "ONTOLOGY", "1:N", ""),
]


def _build_relation(wb: Workbook) -> None:
    ws = wb.create_sheet("术语关系")
    n = len(_REL_COLS)
    _notice_row(ws, 1,
        "【术语关系】* 为必填项。\n"
        "源术语编码 / 目标术语编码：优先填《术语》Sheet 中的【术语编码】；"
        "如该术语编码留空（新增时系统生成），可填【术语名称】代替，系统导入时自动解析。\n"
        "关系类别：ONTOLOGY=本体结构关系（如包含、归属），BUSINESS=业务语义关系（如负责、创建）。",
        n, height=52)
    _header_row(ws, 2, [(d, r, w) for d, r, w, _ in _REL_COLS])
    _desc_row(ws, 3, [desc for _, _, _, desc in _REL_COLS])
    _example_rows(ws, 4, _REL_EXAMPLES)

    _dv_list(ws, '"ONTOLOGY,BUSINESS"', "D4:D1000", allow_blank=False,
             err_msg="只能填 ONTOLOGY 或 BUSINESS", err_title="格式错误")
    _dv_list(ws, '"1:1,1:N,N:1,N:N"', "E4:E1000")

    ws.freeze_panes = "A4"


# ─────────────────────────────────────────────────────────────────────────────
# Sheet6：自定义文件夹说明
# ─────────────────────────────────────────────────────────────────────────────
_README = [
    ("【自定义文件夹 ontology/】使用说明", True, False),
    ("", False, False),
    ("与本 Excel 同级放置 ontology/ 目录，按术语类型分子文件夹存放 JSON 定义文件：", False, False),
    ("", False, False),
    ("  ontology/", False, True),
    ("  ├── objects/      ← OBJ 对象术语的 JSON 文件", False, True),
    ("  ├── views/        ← VIEW 视图术语的 JSON 文件", False, True),
    ("  ├── actions/      ← ACTION 动作术语的 JSON 文件", False, True),
    ("  └── functions/    ← FUNC 函数术语的 JSON 文件", False, True),
    ("", False, False),
    ("《术语》Sheet 的【本体定义文件路径】列填写相对于 ontology/ 的子路径，例如：", False, False),
    ("  objects/sales_customer.json", False, True),
    ("  views/scene_01_data_analysis.json", False, True),
    ("", False, False),
    ("对象术语 JSON 结构（objects/xxx.json）：", True, False),
    ("""{
  "object_code": "sales_customer",
  "object_name": "客户对象",
  "description": "CRM客户表",
  "source_type": "DB",
  "fields": [
    { "field_code": "id",           "field_name": "主键ID",   "field_type": "BIGINT", "is_primary_key": true },
    { "field_code": "customerName", "field_name": "客户名称", "field_type": "STRING" }
  ],
  "relations": []
}""", False, True),
    ("", False, False),
    ("视图术语 JSON 结构（views/xxx.json）：", True, False),
    ("""{
  "view_id":     "scene_01_data_analysis",
  "view_name":   "在线查数分析场景",
  "description": "销售数据分析场景",
  "object_ids":  ["sales_customer", "po_users"],
  "relations": [
    {
      "relation_code": "rel_001",
      "relation_name": "人员归属组织",
      "source_class":  "po_users",
      "target_class":  "po_organization",
      "relation_type": "MANY_TO_ONE"
    }
  ]
}""", False, True),
]


def _build_readme(wb: Workbook) -> None:
    ws = wb.create_sheet("📁 自定义文件夹说明")
    _col_w(ws, 1, 95)
    for row_idx, (line, is_title, is_code) in enumerate(_README, 1):
        cell = ws.cell(row=row_idx, column=1, value=line)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if is_title:
            cell.font = Font(bold=True, size=11, color="1F497D")
        elif is_code:
            cell.font = Font(size=10, name="Courier New")
        else:
            cell.font = Font(size=10)
        lines_in_cell = line.count("\n") + 1
        ws.row_dimensions[row_idx].height = max(15, lines_in_cell * 15)


# ─────────────────────────────────────────────────────────────────────────────
# 主入口
# ─────────────────────────────────────────────────────────────────────────────
def main() -> None:
    """生成导入模板 Excel 文件。"""
    wb = Workbook()
    # 删掉默认的空 Sheet
    wb.remove(wb.active)

    _build_domain(wb)
    _build_library(wb)
    _build_term_type(wb)
    _build_term(wb)
    _build_relation(wb)
    _build_readme(wb)

    out = Path(__file__).parent / "知识构建模板.xlsx"
    wb.save(out)
    print(f"已生成：{out}")


if __name__ == "__main__":
    main()
